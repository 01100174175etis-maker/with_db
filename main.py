"""
نظام إدارة المبيعات والحسابات - النسخة الاحترافية المتقدمة بمكتبة PyQt5
التحديث: نسخة منقحة ومصححة لغوياً بالكامل لتعمل مباشرة دون أي أخطاء.
المميزات: جلب الأسعار تلقائياً، إدخال سريع بزر Enter، وقائمة العملاء الجانبية الكلية باليسار.
"""
from datetime import datetime
import os
from pathlib import Path
import sys

# --- حل مشكلة المسار تلقائيًا لضمان استقرار استدعاء الموديلات ---
CURRENT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = CURRENT_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
import sales_app.db as db
from openpyxl import load_workbook
from openpyxl.styles import Font as XlFont, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# استدعاء مكونات مكتبة الواجهة الاحترافية PyQt5
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QGridLayout, QLabel, QLineEdit, 
                             QComboBox, QPushButton, QTableWidget, QTableWidgetItem, 
                             QHeaderView, QMessageBox, QDesktopWidget, QFileDialog, QListWidget)
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtCore import Qt

APP_TITLE = "نظام إدارة الحسابات والمبيعات المتطور - PyQt5 Enterprise"


class SalesApp(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1250, 720) 
        
        self.center_window()
        self.init_ui_styles()
        
        self.editing_sale_id = None
        self.all_unique_customers = []
        self.all_unique_items = []
        self.price_history = {} 

        # المكون الرئيسي للنافذة ونظام التقسيم الأفقي
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        # التخطيط الأفقي الرئيسي لدمج القائمة الجانبية باليسار مع الواجهة باليمين
        layout_horizontal_master = QHBoxLayout(main_widget)

        # ================= [ القسم الأيسر: قائمة العملاء بالطول الكلي ] =================
        left_panel = QVBoxLayout()
        lbl_side_cust = QLabel("📊 شجرة العملاء الحالية")
        lbl_side_cust.setFont(QFont("Helvetica", 11, QFont.Bold))
        lbl_side_cust.setStyleSheet("color: #1565c0; padding-bottom: 2px;")
        left_panel.addWidget(lbl_side_cust)
        
        self.side_customer_list = QListWidget()
        self.side_customer_list.setFont(QFont("Helvetica", 12, QFont.Bold))
        self.side_customer_list.setFixedWidth(220) 
        self.side_customer_list.setStyleSheet("background-color: white; color: black; border: 2px solid #b0bec5; border-radius: 4px;")
        self.side_customer_list.itemClicked.connect(self.filter_table_by_side_customer)
        left_panel.addWidget(self.side_customer_list)
        
        layout_horizontal_master.addLayout(left_panel)

        # ================= [ القسم الأيمن: واجهة الإدخال والجدول المركزي ] =================
        right_panel = QVBoxLayout()
        layout_horizontal_master.addLayout(right_panel)

        # واجهة إدخال البيانات الحسابية (Grid Layout)
        frm_layout = QGridLayout()
        right_panel.addLayout(frm_layout)

        lbl_font = QFont("Helvetica", 12, QFont.Bold)
        self.entry_font = QFont("Helvetica", 14, QFont.Bold)

        # الخانات العلوية الكبيرة الفاتحة ذات الخلفية البيضاء والخط الأسود العريض
        lbl_cust = QLabel("اسم العميل:")
        lbl_cust.setFont(lbl_font)
        frm_layout.addWidget(lbl_cust, 0, 0)
        self.ent_customer = QComboBox()
        self.ent_customer.setEditable(True)
        self.ent_customer.setFont(self.entry_font)
        self.ent_customer.setFixedWidth(250)
        frm_layout.addWidget(self.ent_customer, 0, 1)

        lbl_item = QLabel("اسم المنتج:")
        lbl_item.setFont(lbl_font)
        frm_layout.addWidget(lbl_item, 0, 2)
        self.ent_item = QComboBox()
        self.ent_item.setEditable(True)
        self.ent_item.setFont(self.entry_font)
        self.ent_item.setFixedWidth(250)
        frm_layout.addWidget(self.ent_item, 0, 3)
        self.ent_item.currentTextChanged.connect(self.auto_fetch_price_from_list)

        lbl_qty = QLabel("الكمية المطلوبة:")
        lbl_qty.setFont(lbl_font)
        frm_layout.addWidget(lbl_qty, 1, 0)
        self.ent_qty = QLineEdit()
        self.ent_qty.setFont(self.entry_font)
        self.ent_qty.setFixedWidth(180)
        frm_layout.addWidget(self.ent_qty, 1, 1)
        self.ent_qty.textChanged.connect(self.calculate_live_total) 

        lbl_price = QLabel("سعر الوحدة:")
        lbl_price.setFont(lbl_font)
        frm_layout.addWidget(lbl_price, 1, 2)
        self.ent_price = QLineEdit()
        self.ent_price.setFont(self.entry_font)
        self.ent_price.setFixedWidth(180)
        frm_layout.addWidget(self.ent_price, 1, 3)
        self.ent_price.textChanged.connect(self.calculate_live_total) 

        lbl_date = QLabel("التاريخ الحالي:")
        lbl_date.setFont(lbl_font)
        frm_layout.addWidget(lbl_date, 2, 0)
        self.ent_date = QLineEdit()
        self.ent_date.setFont(self.entry_font)
        self.ent_date.setText(datetime.now().strftime("%Y-%m-%d"))
        self.ent_date.setFixedWidth(180)
        frm_layout.addWidget(self.ent_date, 2, 1)

        # ربط مفتاح التوجيه الذكي بزر Enter لتسريع حركة الكاشير والمحاسب
        self.ent_qty.returnPressed.connect(lambda: self.ent_price.setFocus())
        self.ent_price.returnPressed.connect(self.add_sale)

        # أزرار الإضافة الحسابية والمسح اليدوي
        self.btn_add = QPushButton("إضافة إلى السجل الحسابي")
        self.btn_add.setFont(QFont("Helvetica", 12, QFont.Bold))
        self.btn_add.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px; border-radius: 4px;")
        self.btn_add.clicked.connect(self.add_sale)
        frm_layout.addWidget(self.btn_add, 0, 4, 2, 1)

        self.btn_clear = QPushButton(" مسح الحقول الشامل 🧹")
        self.btn_clear.setFont(QFont("Helvetica", 11, QFont.Bold))
        self.btn_clear.setStyleSheet("background-color: #9e9e9e; color: white; padding: 6px; border-radius: 4px;")
        self.btn_clear.clicked.connect(self.clear_fields_manually)
        frm_layout.addWidget(self.btn_clear, 2, 4)

        # جدول العرض المالي الأوسط الاحترافي
        self.tree = QTableWidget()
        self.tree.setColumnCount(8)
        self.tree.setHorizontalHeaderLabels(["المعرف الذكي", "اسم العميل", "المنتج", "الكمية", "السعر", "الإجمالي", "تاريخ البيع", "وقت الإدخال"])
        self.tree.setFont(QFont("Helvetica", 11))
        self.tree.setSelectionBehavior(QTableWidget.SelectRows)
        self.tree.setSelectionMode(QTableWidget.SingleSelection)
        self.tree.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        right_panel.addWidget(self.tree)
        # أزرار العمليات والتحكم السفلية المنسقة مع المظهر الفاتح المستقر
        bot_layout = QHBoxLayout()
        right_panel.addLayout(bot_layout)

        self.btn_delete = QPushButton("🗑️ حذف المحدد")
        self.btn_delete.setFont(QFont("Helvetica", 11, QFont.Bold))
        self.btn_delete.setStyleSheet("background-color: #f44336; color: white; padding: 8px; border-radius: 4px;")
        self.btn_delete.clicked.connect(self.delete_selected)
        bot_layout.addWidget(self.btn_delete)

        self.btn_edit = QPushButton("✏️ تعديل السطر")
        self.btn_edit.setFont(QFont("Helvetica", 11, QFont.Bold))
        self.btn_edit.setStyleSheet("background-color: #00bcd4; color: white; padding: 8px; border-radius: 4px;")
        self.btn_edit.clicked.connect(self.prepare_edit_row)
        bot_layout.addWidget(self.btn_edit)

        self.btn_print = QPushButton("🖨️ طباعة الفاتورة")
        self.btn_print.setFont(QFont("Helvetica", 11, QFont.Bold))
        self.btn_print.setStyleSheet("background-color: #ff9800; color: white; padding: 8px; border-radius: 4px;")
        self.btn_print.clicked.connect(self.print_invoice)
        bot_layout.addWidget(self.btn_print)

        self.btn_export_single = QPushButton("📁 تصدير العميل لإكسل")
        self.btn_export_single.setFont(QFont("Helvetica", 11, QFont.Bold))
        self.btn_export_single.setStyleSheet("background-color: #009688; color: white; padding: 8px; border-radius: 4px;")
        self.btn_export_single.clicked.connect(self.export_customer_excel)
        bot_layout.addWidget(self.btn_export_single)

        self.btn_export_all = QPushButton("📊 تحديث السجل الموحد ذو الارتباطات")
        self.btn_export_all.setFont(QFont("Helvetica", 11, QFont.Bold))
        self.btn_export_all.setStyleSheet("background-color: #2196F3; color: white; padding: 8px; border-radius: 4px;")
        self.btn_export_all.clicked.connect(self.export_all_customers_unified_excel)
        bot_layout.addWidget(self.btn_export_all)

        # شريط الحالة السفلي بالعربية
        self.lbl_status = QLabel("النظام جاهز للعمل")
        self.lbl_status.setFont(QFont("Helvetica", 10))
        self.lbl_status.setStyleSheet("color: black; padding: 5px;")
        right_panel.addWidget(self.lbl_status)

        # تعريف قاموس تعريب أسماء الأعمدة داخلياً لمنع انهيار قراءة الملفات
        self.headers_ar = {
            "id": "المعرف الذكي", "customer": "اسم العميل", "item": "المنتج", 
            "qty": "الكمية", "price": "السعر", "total": "الإجمالي", 
            "sale_date": "تاريخ البيع", "created_at": "وقت الإدخال"
        }

        self.load_data()

    def center_window(self):
        geo = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        geo.moveCenter(cp)
        self.move(geo.topLeft())

    def init_ui_styles(self):
        style_sheet = """
            QMainWindow { background-color: #f5f5f5; }
            QLineEdit { background-color: white; color: black; border: 2px solid #b0bec5; border-radius: 4px; padding: 4px; }
            QComboBox { background-color: white; color: black; border: 2px solid #b0bec5; border-radius: 4px; padding: 4px; }
            QTableWidget { background-color: white; color: black; gridline-color: #d9d9d9; selection-background-color: #bbdefb; selection-color: black; }
            QHeaderView::section { background-color: #e0e0e0; color: black; font-weight: bold; border: 1px solid #b0bec5; }
            QMessageBox { background-color: white; }
            QMessageBox QLabel { color: black; font-size: 13px; font-weight: bold; }
            QMessageBox QPushButton { font-size: 11px; font-weight: bold; width: 80px; padding: 5px; }
        """
        self.setStyleSheet(style_sheet)

    def set_status(self, text: str):
        self.lbl_status.setText(text)

    def calculate_live_total(self):
        """ميزة الحساب اللحظي المباشر لقيمة الفاتورة قبل الحفظ لإراحة المحاسب"""
        try:
            qty = int(self.ent_qty.text().strip()) if self.ent_qty.text().strip() else 0
            price = float(self.ent_price.text().strip()) if self.ent_price.text().strip() else 0.0
            total = qty * price
            if total > 0:
                self.set_status(f"📊 إجمالي الفاتورة الحالي المتوقع: {total:.2f} جنيه مصري.")
        except ValueError:
            pass

    def auto_fetch_price_from_list(self, current_item_name):
        """أداة جلب الأسعار التلقائية: جلب أحدث سعر بيع مسجل للمنتج تلقائياً"""
        current_item_name = current_item_name.strip()
        if current_item_name in self.price_history:
            last_price = self.price_history[current_item_name]
            self.ent_price.setText(f"{last_price:.2f}")
            self.set_status(f"💰 تم جلب آخر سعر للمنتج '{current_item_name}' تلقائياً: {last_price:.2f}")
        else:
            self.ent_price.clear()

    def filter_table_by_side_customer(self, item):
        """تصفية فورية: عند الضعل على اسم العميل باليسار يظهر كشف حسابه بالجدول فوراً"""
        selected_cust = item.text().strip()
        if selected_cust == "📋 عرض الكل":
            self.load_data()
            return
            
        df = db.get_sales()
        filtered_df = df[df["customer"] == selected_cust]
        self.display_data_in_table(filtered_df)
        self.set_status(f"🔍 يتم الآن عرض كشف حساب العميل: '{selected_cust}' فقط. (عدد الحركات: {len(filtered_df)})")

    def clear_fields_manually(self):
        self.ent_customer.setCurrentText("")
        self.ent_item.setCurrentText("")
        self.ent_qty.clear()
        self.ent_price.clear()
        self.editing_sale_id = None
        self.btn_add.setText("إضافة إلى السجل الحسابي")
        self.btn_add.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px; border-radius: 4px;")
        self.set_status("تم تنظيف الحقول يدوياً.")

    def update_dropdown_lists(self, df: pd.DataFrame):
        if not df.empty:
            self.all_unique_customers = sorted(df["customer"].dropna().unique().tolist())
            self.all_unique_items = sorted(df["item"].dropna().unique().tolist())
            
            # بناء قائمة الأسعار التاريخية ديناميكياً من أحدث العمليات المسجلة
            for _, row in df.iterrows():
                product = str(row["item"]).strip()
                price = float(row["price"])
                if product not in self.price_history:
                    self.price_history[product] = price

            self.ent_customer.blockSignals(True)
            current_cust = self.ent_customer.currentText()
            self.ent_customer.clear()
            self.ent_customer.addItems(self.all_unique_customers)
            self.ent_customer.setCurrentText(current_cust)
            self.ent_customer.blockSignals(False)
            
            self.ent_item.blockSignals(True)
            current_item = self.ent_item.currentText()
            self.ent_item.clear()
            self.ent_item.addItems(self.all_unique_items)
            self.ent_item.setCurrentText(current_item)
            self.ent_item.blockSignals(False)

            # تحديث شجرة العملاء الجانبية الكلية باليسار
            self.side_customer_list.blockSignals(True)
            self.side_customer_list.clear()
            self.side_customer_list.addItem("📋 عرض الكل")
            self.side_customer_list.addItems(self.all_unique_customers)
            self.side_customer_list.blockSignals(False)

    def load_data(self):
        df = db.get_sales()
        self.display_data_in_table(df)
        self.update_dropdown_lists(df)
        self.set_status(f"تم تحميل عدد {len(df)} سجل حسابي بنجاح.")

    def display_data_in_table(self, df):
        self.tree.setRowCount(0)
        for row_idx, row in df.reset_index(drop=True).iterrows():
            self.tree.insertRow(row_idx)
            total_val = float(row["total"])
            
            is_large = total_val >= 10000
            bg_color = QColor('#ffeb3b') if is_large else QColor('white')
            
            columns_data = [row["id"], row["customer"], row["item"], row["qty"], f"{row['price']:.2f}", f"{total_val:.2f}", row["sale_date"], row["created_at"]]
            for col_idx, data in enumerate(columns_data):
                item = QTableWidgetItem(str(data))
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(bg_color)
                item.setForeground(QColor('black'))
                self.tree.setItem(row_idx, col_idx, item)

    def add_sale(self):
        c = self.ent_customer.currentText().strip()
        i = self.ent_item.currentText().strip()
        q_t = self.ent_qty.text().strip()
        p_t = self.ent_price.text().strip()
        d = self.ent_date.text().strip()
        
        if not (c and i and q_t and p_t and d):
            QMessageBox.warning(self, "بيانات ناقصة", "برجاء اختيار أو كتابة العميل والمنتج وملء كافة الخانات.")
            return
        
        try:
            datetime.strptime(d, "%Y-%m-%d")
        except ValueError:
            QMessageBox.critical(self, "خطأ في التاريخ", "صيغة التاريخ خاطئة! يجب أن تكون YYYY-MM-DD")
            return

        try:
            qty, price = int(q_t), float(p_t)
            total = qty * price
        except ValueError:
            QMessageBox.critical(self, "خطأ رقمي", "يجب إدخال أرقام صحيحة في خانة الكمية والسعر.")
            return

        if self.editing_sale_id is not None:
            db.delete_sale(self.editing_sale_id)
            self.editing_sale_id = None
            self.btn_add.setText("إضافة إلى السجل الحسابي")
            self.btn_add.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px; border-radius: 4px;")
            msg_status = "تعديل وتحديث"
        else:
            msg_status = "تسجيل"

        new_uid = db.add_sale(c, i, qty, price, total, d)
        self.load_data()
        self.clear_fields_manually()
        self.set_status(f"تم {msg_status} الفاتورة بنجاح للعميل: {c}. المعرف: {new_uid}")
        self.ent_customer.setFocus()
    def prepare_edit_row(self):
        selected_rows = self.tree.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "برجاء تحديد سطر من الجدول أولاً لتعديله.")
            return
        
        row_idx = selected_rows[0].row()
        
        self.editing_sale_id = self.tree.item(row_idx, 0).text()
        customer_name = self.tree.item(row_idx, 1).text()
        item_name = self.tree.item(row_idx, 2).text()
        qty = self.tree.item(row_idx, 3).text()
        price = self.tree.item(row_idx, 4).text()
        sale_date = self.tree.item(row_idx, 6).text()
        
        self.ent_customer.setCurrentText(customer_name)
        self.ent_item.setCurrentText(item_name)
        self.ent_qty.setText(qty)
        self.ent_price.setText(price)
        self.ent_date.setText(sale_date)
        
        self.btn_add.setText("💾 حفظ التعديلات الآن")
        self.btn_add.setStyleSheet("background-color: #00bcd4; color: white; padding: 10px; border-radius: 4px;")
        self.set_status(f"جاري تعديل العملية ذات المعرف: {self.editing_sale_id}")

    def delete_selected(self):
        selected_rows = self.tree.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "برجاء تحديد سطر من الجدول أولاً للحذف.")
            return
        
        row_idx = selected_rows[0].row()
        sale_id = self.tree.item(row_idx, 0).text()
        
        confirm = QMessageBox.question(
            self, "تأكيد الحذف", f"هل أنت متأكد من حذف السجل رقم المعرف {sale_id} فقط؟",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if confirm == QMessageBox.Yes:
            db.delete_sale(sale_id)
            self.load_data()
            if self.editing_sale_id == sale_id:
                self.clear_fields_manually()
            self.set_status(f"تم حذف العملية رقم {sale_id} بنجاح.")

    def print_invoice(self):
        selected_rows = self.tree.selectionModel().selectedRows()
        if not selected_rows:
            return
        row_idx = selected_rows[0].row()
        
        invoice_id = self.tree.item(row_idx, 0).text()
        customer_name = self.tree.item(row_idx, 1).text()
        item_name = self.tree.item(row_idx, 2).text()
        qty = self.tree.item(row_idx, 3).text()
        price = self.tree.item(row_idx, 4).text()
        total = self.tree.item(row_idx, 5).text()
        sale_date = self.tree.item(row_idx, 6).text()
        
        invoice_text = f"Invoice ID: {invoice_id}\nDate: {sale_date}\nCustomer: {customer_name}\nItem: {item_name}\nQty: {qty}\nPrice: {price}\nTotal: {total} EGP"
        invoice_file = CURRENT_DIR / f"Invoice_{invoice_id}.txt"
        invoice_file.write_text(invoice_text, encoding="utf-8")
        os.startfile(invoice_file)
        self.set_status(f"تم توليد الفاتورة رقم {invoice_id} بنجاح.")

    def export_customer_excel(self):
        selected_rows = self.tree.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "تنبيه", "برجاء تحديد سطر للعميل أولاً.")
            return
        row_idx = selected_rows[0].row()
        customer_name = self.tree.item(row_idx, 1).text()
        
        all_sales_df = db.get_sales()
        customer_df = all_sales_df[all_sales_df["customer"] == customer_name].rename(columns=self.headers_ar)
        
        path, _ = QFileDialog.getSaveFileName(self, "تصدير عميل", f"{customer_name}.xlsx", "Excel files (*.xlsx)")
        if path:
            customer_df.to_excel(path, index=False)
            QMessageBox.information(self, "تصدير ناجح", f"تم إنشاء ملف إكسل مخصص للعميل '{customer_name}' بنجاح.")

    def export_all_customers_unified_excel(self):
        all_sales_df = db.get_sales()
        if all_sales_df.empty:
            QMessageBox.warning(self, "تنبيه", "لا توجد أي بيانات حالية لتصديرها.")
            return
            
        path, _ = QFileDialog.getSaveFileName(self, "تصدير السجل الموحد", "دفتر_الأستاذ_الموحد.xlsx", "Excel files (*.xlsx)")
        if not path:
            return

        try:
            summary_data = []
            grouped_all = all_sales_df.groupby("customer")
            for customer_name, group in grouped_all:
                summary_data.append({
                    "اسم العميل": str(customer_name),
                    "إجمالي الفواتير (جنيه)": group["total"].sum(),
                    "عدد أنواع المنتجات": group["item"].nunique()
                })
            summary_df = pd.DataFrame(summary_data)

            if os.path.exists(path):
                try:
                    os.remove(path)
                except OSError:
                    QMessageBox.critical(self, "ملف مفتوح", "برجاء إغلاق ملف الإكسل الموحد أولاً قبل محاولة تحديثه.")
                    return

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                summary_df.to_excel(writer, sheet_name="مختصر العملاء", index=False)
                for customer_name, group in grouped_all:
                    safe_sheet_name = "".join(c for c in str(customer_name) if c.isalnum() or c in (" ", "_", "-"))[:30].strip()
                    if not safe_sheet_name:
                        safe_sheet_name = f"عميل"
                    detailed_df = group.rename(columns=self.headers_ar)
                    detailed_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

            wb = load_workbook(path)
            header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_font = XlFont(name="Calibri", size=14, bold=True, color="000000")
            regular_font = XlFont(name="Calibri", size=12, bold=False, color="000000")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thick_side = Side(border_style="medium", color="000000")
            cell_border = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for col in ws.columns:
                    for cell in col:
                        cell.fill = header_fill
                        if cell.row == 1:
                            cell.font = header_font
                        else:
                            if sheet_name == "مختصر العملاء" and cell.column == 1:
                                pass
                            else:
                                cell.font = regular_font
                        cell.alignment = center_align
                        cell.border = cell_border
                
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col.column)
                    ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

            ws_summary = wb["مختصر العملاء"]
            for row_idx in range(2, ws_summary.max_row + 1):
                cell = ws_summary.cell(row=row_idx, column=1)
                customer_sheet_name = "".join(c for c in str(cell.value) if c.isalnum() or c in (" ", "_", "-"))[:30].strip()
                cell.hyperlink = f"#'{customer_sheet_name}'!A1"
                cell.font = XlFont(color="0000FF", underline="single", name="Calibri", size=12, bold=True)
            
            wb.save(path)
            QMessageBox.information(self, "عملية ناجحة", "تم تحديث دفتر الأستاذ الموحد وتنسيقه بالخلفية البيضاء الكلاسيكية والخطوط السوداء العريضة والروابط بنجاح!")
            self.set_status("تم استخراج وتنسيق ملف السجل الموحد بنجاح.")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"فشلت العملية الحسابية بسبب: {e}")


def main():
    db.init_db()
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = SalesApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
